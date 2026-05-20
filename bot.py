import os
import json
import logging
import io
import base64
import threading
import re
from http.server import HTTPServer, BaseHTTPRequestHandler

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    filters, ConversationHandler, ContextTypes, CallbackQueryHandler
)
import gspread
from google.oauth2.service_account import Credentials as SACredentials
from googleapiclient.discovery import build as gdrive_build
from googleapiclient.http import MediaIoBaseUpload
import openpyxl
import xlrd

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

BOT_TOKEN      = os.environ['BOT_TOKEN']
SHEET_ID       = os.environ['SHEET_ID']
CHANNEL_ID     = int(os.environ.get('CHANNEL_ID', '0'))
DRIVE_FOLDER_ID = os.environ.get('DRIVE_FOLDER_ID', '')

HEADER = ['Дата', 'Поставщик', 'Бренд', 'Товар', 'Цена KRW', 'Цена USD']
DEFAULT_RATE = 1450

# ── Conversation states ────────────────────────────────────────────────────────
WAITING_SUPPLIER  = 1
WAITING_PRICE_COL = 2
WAITING_NAME_COL  = 3
WAITING_RATE      = 4
WAITING_BRAND     = 5

# ── Column detection keywords ──────────────────────────────────────────────────
PRICE_KEYWORDS = [
    'supply price', 'supply', '공급가', '납품가', '공급단가', '공급가격',
    '단가', '원가', '도매가', '도매단가', 'wholesale', 'cost', 'unit price',
    '가격', '공급', '매입가', '매입단가', '원단가', 'price'
]
NAME_KEYWORDS = [
    'product name', 'product', '상품명', '품명', '제품명', '상품',
    'item', 'name', '품목', '아이템', '모델명', '모델', '제품', '항목'
]
EXCLUDE_PRICE  = ['retail', 'msrp', 'recommend', 'consumer', '소비자', '판매가', '소매',
                  'rate', 'ratio', '비율', '할인율', '할인', '세율', 'discount', 'margin']
VOLUME_KEYWORDS = [
    '용량', '내용량', 'volume', 'capacity', 'size', '사이즈', '규격', '중량',
    'weight', 'ml', 'g(', '(g)', 'oz', 'liter'
]
QTY_KEYWORDS = [
    '박스수량', '입수', 'box qty', 'pcs/box', 'qty/box', '개입', '박스당',
    'units/box', 'pieces/box', 'ea/box', 'per box'
]


# ── Health server ──────────────────────────────────────────────────────────────
class HealthHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200); self.end_headers(); self.wfile.write(b'OK')
    def log_message(self, *a): pass


# ── Google credentials ─────────────────────────────────────────────────────────
def _creds_dict():
    raw = os.environ['GOOGLE_CREDENTIALS']
    try:
        return json.loads(raw)
    except (json.JSONDecodeError, ValueError):
        return json.loads(base64.b64decode(raw).decode('utf-8'))


def get_gspread_client():
    return gspread.service_account_from_dict(
        _creds_dict(),
        scopes=['https://www.googleapis.com/auth/spreadsheets',
                'https://www.googleapis.com/auth/drive']
    )


def get_drive_service():
    creds = SACredentials.from_service_account_info(
        _creds_dict(),
        scopes=['https://www.googleapis.com/auth/drive']
    )
    return gdrive_build('drive', 'v3', credentials=creds, cache_discovery=False)


# ── Google Drive helpers ───────────────────────────────────────────────────────
def _get_or_create_folder(service, name, parent_id):
    safe = name[:40].replace("'", "\\'")
    q = (f"name='{safe}' and '{parent_id}' in parents "
         f"and mimeType='application/vnd.google-apps.folder' and trashed=false")
    res = service.files().list(q=q, fields='files(id)').execute()
    files = res.get('files', [])
    if files:
        return files[0]['id']
    folder = service.files().create(
        body={'name': name[:40], 'mimeType': 'application/vnd.google-apps.folder',
              'parents': [parent_id]},
        fields='id'
    ).execute()
    return folder['id']


def upload_to_drive(file_bytes: bytes, filename: str, supplier: str, date_str: str):
    """Upload original Excel to Drive: ROOT / Supplier / date_filename.xlsx"""
    if not DRIVE_FOLDER_ID:
        return None
    try:
        service = get_drive_service()
        supplier_folder = _get_or_create_folder(service, supplier, DRIVE_FOLDER_ID)
        safe_name = f'{date_str}_{filename}'
        media = MediaIoBaseUpload(
            io.BytesIO(file_bytes),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resumable=False
        )
        result = service.files().create(
            body={'name': safe_name, 'parents': [supplier_folder]},
            media_body=media,
            fields='id,webViewLink'
        ).execute()
        logger.info('Uploaded to Drive: %s', result.get('webViewLink'))
        return result.get('webViewLink', '')
    except Exception as e:
        logger.warning('Drive upload failed: %s', e)
        return None


# ── Google Sheets helpers ──────────────────────────────────────────────────────
def get_base_sheet(spreadsheet):
    sheet = None
    try:
        sheet = spreadsheet.worksheet('База')
    except gspread.exceptions.WorksheetNotFound:
        sheet = spreadsheet.sheet1
        try:
            sheet.update_title('База')
        except Exception:
            pass

    existing = sheet.get_all_values()
    # Migrate very old format (Product, Supplier, Price KRW, Price USD, Updated)
    if existing and existing[0] == ['Product', 'Supplier', 'Price KRW', 'Price USD', 'Updated']:
        logger.info('Migrating very old sheet format...')
        new_data = [HEADER]
        for row in existing[1:]:
            if len(row) >= 5 and row[0] and row[2]:
                new_data.append([row[4], row[1], '', row[0], row[2], row[3]])
        sheet.clear()
        sheet.update(new_data, value_input_option='RAW')
    # Migrate v1 format (Дата, Поставщик, Товар, Цена KRW, Цена USD) — no Brand column
    elif existing and existing[0] == ['Дата', 'Поставщик', 'Товар', 'Цена KRW', 'Цена USD']:
        logger.info('Migrating v1 sheet format (adding Бренд column)...')
        new_data = [HEADER]
        for row in existing[1:]:
            if len(row) >= 4:
                new_data.append([
                    row[0], row[1], '',
                    row[2], row[3],
                    row[4] if len(row) > 4 else ''
                ])
        sheet.clear()
        sheet.update(new_data, value_input_option='RAW')
    elif not existing or existing[0] != HEADER:
        sheet.clear()
        sheet.append_row(HEADER)
    return sheet


def get_latest_prices(all_data):
    """Most recent price per (product, supplier) from append-only history.
    Returns: {(product, supplier): (date, price_krw, brand)}
    """
    latest = {}
    for row in all_data[1:]:
        if len(row) < 5: continue
        # New format: [date, supplier, brand, product, price_krw, price_usd]
        date, supplier, brand, product, price_raw = row[0], row[1], row[2], row[3], row[4]
        if not product or not supplier or not price_raw: continue
        try:
            price_krw = int(float(str(price_raw).replace(',', '')))
        except (ValueError, TypeError):
            continue
        if price_krw <= 0: continue
        key = (product, supplier)
        if key not in latest or date > latest[key][0]:
            latest[key] = (date, price_krw, brand)
    return latest


def save_to_sheet(data_rows, supplier, brand, date_str, rate, spreadsheet):
    sheet = get_base_sheet(spreadsheet)
    new_rows = [
        [date_str, supplier, brand, product, price_krw, round(price_krw / rate, 2)]
        for product, price_krw in data_rows
    ]
    sheet.append_rows(new_rows, value_input_option='RAW')
    for fn_name, fn_args in [
        ('rebuild_comparison',     (spreadsheet, rate)),
        ('rebuild_best_prices',    (spreadsheet, rate)),
        ('rebuild_brand_catalog',  (spreadsheet,)),
        ('rebuild_supplier_catalog', (spreadsheet,)),
        ('rebuild_dashboard',      (spreadsheet,)),
    ]:
        try:
            globals()[fn_name](*fn_args)
        except Exception as e:
            logger.warning('%s failed: %s', fn_name, e)
    return len(new_rows)


def _get_or_create_sheet(spreadsheet, title, rows=100, cols=20):
    try:
        return spreadsheet.worksheet(title)
    except gspread.exceptions.WorksheetNotFound:
        return spreadsheet.add_worksheet(title, rows=rows, cols=cols)


def _load_latest(spreadsheet):
    try:
        base = spreadsheet.worksheet('База')
    except Exception:
        base = spreadsheet.sheet1
    all_data = base.get_all_values()
    if len(all_data) <= 1:
        return None, all_data
    return get_latest_prices(all_data), all_data


def rebuild_comparison(spreadsheet, rate=DEFAULT_RATE):
    latest, all_data = _load_latest(spreadsheet)
    if not latest:
        return
    pivot = {}
    suppliers_set = set()
    brands_map = {}  # product → brand
    for (product, supplier), (_, price_krw, brand) in latest.items():
        pivot.setdefault(product, {})[supplier] = price_krw
        suppliers_set.add(supplier)
        if brand:
            brands_map[product] = brand
    if not pivot:
        return
    suppliers = sorted(suppliers_set)
    comp = _get_or_create_sheet(spreadsheet, 'Сравнение', rows=len(pivot)+10, cols=len(suppliers)+5)
    header = ['Бренд', 'Товар'] + suppliers + ['Лучший поставщик', 'Мин. цена ₩', 'Мин. цена $']
    table = [header]
    for product in sorted(pivot.keys()):
        prices = pivot[product]
        brand = brands_map.get(product, '')
        row = [brand, product] + [prices.get(s, '') for s in suppliers]
        best = min(prices, key=prices.get)
        min_krw = min(prices.values())
        row.extend([best, min_krw, round(min_krw / rate, 2)])
        table.append(row)
    comp.clear()
    comp.update(table, value_input_option='RAW')


def rebuild_best_prices(spreadsheet, rate=DEFAULT_RATE):
    latest, _ = _load_latest(spreadsheet)
    if not latest:
        return
    # Group by product: find min price across all suppliers
    by_product = {}
    for (product, supplier), (date, price_krw, brand) in latest.items():
        entry = by_product.get(product)
        if entry is None or price_krw < entry['min_krw']:
            by_product[product] = {
                'brand': brand, 'min_krw': price_krw,
                'best_supplier': supplier, 'date': date,
            }
        # track all suppliers for this product
        by_product[product].setdefault('all_suppliers', {})
        by_product[product]['all_suppliers'][supplier] = price_krw

    sheet = _get_or_create_sheet(spreadsheet, 'Лучшие цены', rows=len(by_product)+10, cols=8)
    header = ['Бренд', 'Товар', 'Лучший поставщик', 'Мин. цена ₩', 'Мин. цена $', 'Дата', 'Все поставщики (цена ₩)']
    table = [header]
    for product in sorted(by_product.keys(), key=lambda p: (by_product[p].get('brand',''), p)):
        d = by_product[product]
        others = ', '.join(
            f"{s}: {p:,}₩"
            for s, p in sorted(d['all_suppliers'].items(), key=lambda x: x[1])
        )
        table.append([
            d['brand'], product, d['best_supplier'],
            d['min_krw'], round(d['min_krw'] / rate, 2),
            d['date'], others
        ])
    sheet.clear()
    sheet.update(table, value_input_option='RAW')


def rebuild_brand_catalog(spreadsheet):
    latest, _ = _load_latest(spreadsheet)
    if not latest:
        return
    brands = {}  # brand → {suppliers, products, prices}
    for (product, supplier), (date, price_krw, brand) in latest.items():
        key = brand if brand else '(без бренда)'
        b = brands.setdefault(key, {'suppliers': set(), 'products': set(), 'prices': []})
        b['suppliers'].add(supplier)
        b['products'].add(product)
        b['prices'].append(price_krw)

    sheet = _get_or_create_sheet(spreadsheet, 'Бренды', rows=len(brands)+10, cols=6)
    header = ['Бренд', 'Поставщики', 'Кол-во товаров', 'Мин. цена ₩', 'Макс. цена ₩', 'Ср. цена ₩']
    table = [header]
    for brand_name in sorted(brands.keys()):
        b = brands[brand_name]
        prices = b['prices']
        table.append([
            brand_name,
            ', '.join(sorted(b['suppliers'])),
            len(b['products']),
            min(prices),
            max(prices),
            round(sum(prices) / len(prices))
        ])
    sheet.clear()
    sheet.update(table, value_input_option='RAW')


def rebuild_supplier_catalog(spreadsheet):
    latest, _ = _load_latest(spreadsheet)
    if not latest:
        return
    suppliers = {}  # supplier → {brands, products, last_date}
    for (product, supplier), (date, price_krw, brand) in latest.items():
        s = suppliers.setdefault(supplier, {'brands': set(), 'products': set(), 'last_date': ''})
        s['products'].add(product)
        if brand:
            s['brands'].add(brand)
        if date > s['last_date']:
            s['last_date'] = date

    sheet = _get_or_create_sheet(spreadsheet, 'Поставщики', rows=len(suppliers)+10, cols=5)
    header = ['Поставщик', 'Бренды', 'Кол-во товаров', 'Последнее обновление']
    table = [header]
    for supplier_name in sorted(suppliers.keys()):
        s = suppliers[supplier_name]
        table.append([
            supplier_name,
            ', '.join(sorted(s['brands'])) or '—',
            len(s['products']),
            s['last_date']
        ])
    sheet.clear()
    sheet.update(table, value_input_option='RAW')


def rebuild_dashboard(spreadsheet):
    latest, all_data = _load_latest(spreadsheet)
    if latest is None:
        latest = {}

    suppliers = set()
    brands = set()
    products = set()
    last_date = ''
    for (product, supplier), (date, price_krw, brand) in latest.items():
        suppliers.add(supplier)
        products.add(product)
        if brand:
            brands.add(brand)
        if date > last_date:
            last_date = date

    total_records = max(0, len(all_data) - 1)

    sheet = _get_or_create_sheet(spreadsheet, 'Дашборд', rows=15, cols=3)
    table = [
        ['📊 HANWAY TRADE — База прайсов', '', ''],
        ['', '', ''],
        ['Показатель', 'Значение', ''],
        ['Поставщиков', len(suppliers), ''],
        ['Брендов', len(brands), ''],
        ['Уникальных товаров', len(products), ''],
        ['Записей в истории', total_records, ''],
        ['Последнее обновление', last_date or '—', ''],
    ]
    sheet.clear()
    sheet.update(table, value_input_option='RAW')


# ── Excel reading ──────────────────────────────────────────────────────────────
def _find_header_idx(all_rows):
    """Find the most likely header row index (0-based) in the first 25 rows."""
    all_kw = PRICE_KEYWORDS + NAME_KEYWORDS + VOLUME_KEYWORDS + QTY_KEYWORDS
    best_score = -1
    best_idx = 0
    for i, row in enumerate(all_rows[:25]):
        non_empty = sum(1 for v in row if v is not None and str(v).strip())
        if non_empty < 2:
            continue
        h_lower = [str(h).lower().strip().replace('\n', ' ') if h else '' for h in row]
        score = sum(1 for h in h_lower if any(kw in h for kw in all_kw))
        # Bonus: row has many non-empty cells = wider table
        score += non_empty * 0.05
        if score > best_score:
            best_score = score
            best_idx = i
    # Fallback: first row with 3+ non-empty cells
    if best_score <= 0:
        for i, row in enumerate(all_rows[:15]):
            if sum(1 for v in row if v is not None and str(v).strip()) >= 3:
                return i
    return best_idx


def read_excel(file_bytes, file_name):
    ext = file_name.lower().rsplit('.', 1)[-1]
    if ext == 'xls':
        wb = xlrd.open_workbook(file_contents=bytes(file_bytes))
        best = max(range(wb.nsheets), key=lambda i: wb.sheet_by_index(i).nrows * wb.sheet_by_index(i).ncols)
        ws = wb.sheet_by_index(best)
        all_rows = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]
    else:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        best_ws = wb.active
        best_score = 0
        for s in wb.worksheets:
            if s.max_row and s.max_column:
                score = (s.max_row or 0) * (s.max_column or 0)
                if score > best_score:
                    best_score = score
                    best_ws = s
        all_rows = [list(row) for row in best_ws.iter_rows(values_only=True)]
        wb.close()
    if not all_rows:
        return [], []
    header_idx = _find_header_idx(all_rows)
    return all_rows[header_idx], all_rows[header_idx + 1:]


def find_columns(headers, sample_rows=None):
    product_col = price_col = volume_col = qty_col = -1
    h_lower = [str(h).lower().strip().replace('\n', ' ').replace('\r', ' ') if h else '' for h in headers]

    for i, h in enumerate(h_lower):
        for kw in NAME_KEYWORDS:
            if kw in h and ('eng' in h or 'english' in h or '(en' in h):
                product_col = i; break

    if product_col == -1:
        for i, h in enumerate(h_lower):
            for kw in NAME_KEYWORDS:
                if kw in h:
                    product_col = i; break

    def _is_valid_price_col(col_idx, sample_rows):
        """Return True if the column looks like it contains actual prices (not rates/ratios)."""
        if sample_rows is None:
            return True
        vals = []
        for row in sample_rows[:10]:
            if col_idx < len(row) and row[col_idx] is not None:
                try:
                    v = float(str(row[col_idx]).replace(',', ''))
                    vals.append(v)
                except (ValueError, TypeError):
                    pass
        if not vals:
            return True
        avg = sum(vals) / len(vals)
        return avg >= 10  # rates are 0.0-1.0, prices are 1000+ KRW

    for i, h in enumerate(h_lower):
        if any(kw in h for kw in PRICE_KEYWORDS) and not any(ex in h for ex in EXCLUDE_PRICE):
            if _is_valid_price_col(i, sample_rows):
                price_col = i; break

    for i, h in enumerate(h_lower):
        if i in (product_col, price_col): continue
        if any(kw in h for kw in VOLUME_KEYWORDS):
            volume_col = i; break

    for i, h in enumerate(h_lower):
        if i in (product_col, price_col, volume_col): continue
        if any(kw in h for kw in QTY_KEYWORDS):
            qty_col = i; break

    # Fallback: score columns by data type
    if sample_rows and (product_col == -1 or price_col == -1):
        numeric = [0] * len(headers)
        textual = [0] * len(headers)
        for row in sample_rows[:10]:
            for ci, val in enumerate(row):
                if ci >= len(headers) or val is None: continue
                s = str(val).replace(',', '').replace(' ', '').replace('원', '')
                try:
                    float(s); numeric[ci] += 1
                except ValueError:
                    if len(str(val).strip()) > 1: textual[ci] += 1
        if price_col == -1:
            best = max(range(len(numeric)), key=lambda i: numeric[i])
            if numeric[best] > 0: price_col = best
        if product_col == -1:
            best = max(range(len(textual)), key=lambda i: textual[i])
            if textual[best] > 0 and best != price_col: product_col = best

    return product_col, price_col, volume_col, qty_col


# ── Column picker helpers ──────────────────────────────────────────────────────
def _col_items(headers):
    """[(original_index, label), ...] for non-empty headers."""
    return [(i, str(h).strip().replace('\n', ' ')) for i, h in enumerate(headers) if h and str(h).strip()]


def _col_keyboard(headers):
    items = _col_items(headers)
    buttons = []
    row = []
    for num, (col_idx, label) in enumerate(items, 1):
        short = label[:20]
        row.append(InlineKeyboardButton(f'{num}. {short}', callback_data=f'col:{col_idx}'))
        if len(row) == 2:
            buttons.append(row); row = []
    if row:
        buttons.append(row)
    return InlineKeyboardMarkup(buttons), items


def _format_spec(val, header=''):
    s = str(val).strip()
    if not s or s in ('0', '0.0', 'None', ''): return ''
    if any(c.isalpha() for c in s): return s
    for unit in ['ml', 'g', 'oz', 'kg', 'mg', 'l']:
        if unit in header.lower(): return s + unit
    return s


# ── Flow controller ────────────────────────────────────────────────────────────
async def _advance_flow(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Check what info is still missing and ask for it, or proceed to process."""
    ud = context.user_data

    # 1. Supplier
    if not ud.get('supplier'):
        await update.message.reply_text('🏢 Напиши название поставщика:')
        return WAITING_SUPPLIER

    # 2. Brand
    if not ud.get('brand'):
        await update.message.reply_text(
            '🏷️ Напиши название <b>бренда</b> (например: CELIMAX, MEDICUBE, INKOA).\n'
            'Если файл содержит несколько брендов — напиши основной или <code>mixed</code>.',
            parse_mode='HTML'
        )
        return WAITING_BRAND

    headers = ud.get('headers', [])

    # 3. Price column
    if ud.get('price_col') is None:
        keyboard, items = _col_keyboard(headers)
        ud['col_items'] = items
        ud['col_target'] = 'price_col'
        await update.message.reply_text(
            '💰 Не смог автоматически определить столбец с ценой.\n'
            'Выбери какой столбец содержит <b>ЦЕНУ</b>:',
            reply_markup=keyboard, parse_mode='HTML'
        )
        return WAITING_PRICE_COL

    # 4. Product name column
    if ud.get('product_col') is None:
        keyboard, items = _col_keyboard(headers)
        ud['col_items'] = items
        ud['col_target'] = 'product_col'
        await update.message.reply_text(
            '📦 Выбери какой столбец содержит <b>НАЗВАНИЕ ТОВАРА</b>:',
            reply_markup=keyboard, parse_mode='HTML'
        )
        return WAITING_NAME_COL

    # 5. Exchange rate
    if ud.get('rate') is None:
        await update.message.reply_text(
            f'💱 Курс KRW → USD?\n'
            f'Отправь число (например <code>1450</code>) или <code>-</code> чтобы оставить {DEFAULT_RATE}',
            parse_mode='HTML'
        )
        return WAITING_RATE

    # All info ready — process
    return await _do_process(update, context)


async def _col_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle inline keyboard column selection."""
    query = update.callback_query
    await query.answer()

    data = query.data  # 'col:INDEX'
    col_idx = int(data.split(':')[1])
    target = context.user_data.get('col_target')
    if target:
        context.user_data[target] = col_idx

    await query.edit_message_reply_markup(reply_markup=None)
    label = next((lbl for i, lbl in context.user_data.get('col_items', []) if i == col_idx), str(col_idx))
    key_name = '💰 цена' if target == 'price_col' else '📦 название'
    await query.message.reply_text(f'✅ {key_name}: <b>{label}</b>', parse_mode='HTML')

    # Continue flow — fake an update so _advance_flow works
    # We pass the callback query's message as the message
    class _FakeUpdate:
        message = query.message

    return await _advance_flow(_FakeUpdate(), context)


# ── Handlers ───────────────────────────────────────────────────────────────────
async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        '👋 Привет!\n\n'
        '📂 <b>Загрузка прайса:</b> отправь Excel файл поставщика.\n'
        'В подписи можно указать: <code>Поставщик | Бренд</code>\n'
        'или бот спросит сам. Каждая загрузка добавляется в <b>историю с датой</b>.\n\n'
        '🔍 <b>Поиск:</b> напиши название бренда, товара или поставщика.\n'
        'Покажу актуальные цены от всех поставщиков,\n'
        'самая дешёвая отмечена <b>✅</b>.\n\n'
        '📊 <b>Google Sheets листы:</b>\n'
        '  • <b>База</b> — полная история всех прайсов\n'
        '  • <b>Сравнение</b> — товары × поставщики\n'
        '  • <b>Лучшие цены</b> — минимальная цена по каждому товару\n'
        '  • <b>Бренды</b> — каталог брендов и их поставщики\n'
        '  • <b>Поставщики</b> — каталог поставщиков и их бренды\n'
        '  • <b>Дашборд</b> — общая статистика\n\n'
        '☁️ Файлы сохраняются в Google Drive по папкам поставщиков.',
        parse_mode='HTML'
    )


async def cmd_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text('❌ Отменено.')
    return ConversationHandler.END


async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    doc = update.message.document
    if not doc.file_name.lower().endswith(('.xlsx', '.xls')):
        await update.message.reply_text('❌ Нужен Excel файл (.xlsx или .xls)')
        return ConversationHandler.END

    # Download and parse immediately
    dl_msg = await update.message.reply_text('📥 Загружаю файл…')
    try:
        tg_file = await context.bot.get_file(doc.file_id)
        file_bytes = bytes(await tg_file.download_as_bytearray())
        headers, all_rows = read_excel(file_bytes, doc.file_name)
    except Exception as e:
        await dl_msg.edit_text(f'❌ Ошибка загрузки: {e}')
        return ConversationHandler.END
    await dl_msg.delete()

    if not headers:
        await update.message.reply_text('❌ Файл пустой или не читается.')
        return ConversationHandler.END

    # Auto-detect columns
    product_col, price_col, volume_col, qty_col = find_columns(headers, all_rows)

    # Store everything in user_data
    context.user_data.update({
        'file_bytes': file_bytes,
        'file_name': doc.file_name,
        'file_id': doc.file_id,
        'headers': headers,
        'all_rows': all_rows,
        'product_col': product_col if product_col != -1 else None,
        'price_col':   price_col   if price_col != -1   else None,
        'volume_col':  volume_col,
        'qty_col':     qty_col,
        'rate':        None,       # always ask
    })

    # Supplier and brand from caption (format: "Supplier | Brand" or just "Supplier")
    caption = (update.message.caption or '').strip()
    if caption:
        parts = [p.strip() for p in caption.split('|')]
        context.user_data['supplier'] = parts[0] if parts[0] else None
        context.user_data['brand'] = parts[1] if len(parts) > 1 and parts[1] else None
    else:
        context.user_data['supplier'] = None
        context.user_data['brand'] = None

    # Show what was detected
    det_price   = f'«{str(headers[price_col]).strip()}»'   if price_col   != -1 else '❓ не найден'
    det_product = f'«{str(headers[product_col]).strip()}»' if product_col != -1 else '❓ не найден'
    await update.message.reply_text(
        f'📄 <b>{doc.file_name}</b>\n'
        f'Столбцов: {len([h for h in headers if h])}\n'
        f'Строк данных: {len(all_rows)}\n\n'
        f'🔍 Авто-определение:\n'
        f'  💰 Цена: {det_price}\n'
        f'  📦 Товар: {det_product}',
        parse_mode='HTML'
    )

    return await _advance_flow(update, context)


async def handle_supplier_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['supplier'] = update.message.text.strip()
    return await _advance_flow(update, context)


async def handle_brand_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['brand'] = update.message.text.strip()
    return await _advance_flow(update, context)


async def handle_rate_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if text == '-' or text.lower() in ('ок', 'ok', 'да', 'default'):
        context.user_data['rate'] = DEFAULT_RATE
    else:
        try:
            rate = int(float(text.replace(',', '.')))
            if rate < 100 or rate > 10000:
                await update.message.reply_text('Введи реальный курс (например 1450):')
                return WAITING_RATE
            context.user_data['rate'] = rate
        except ValueError:
            await update.message.reply_text(f'Не понял. Введи число или «-» для {DEFAULT_RATE}:')
            return WAITING_RATE
    return await _advance_flow(update, context)


# ── Core processing ────────────────────────────────────────────────────────────
def _esc(s):
    return str(s).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')


async def _do_process(update, context: ContextTypes.DEFAULT_TYPE):
    ud = context.user_data
    file_bytes  = ud['file_bytes']
    file_name   = ud['file_name']
    file_id     = ud['file_id']
    supplier    = ud['supplier']
    brand       = ud.get('brand', '')
    headers     = ud['headers']
    all_rows    = ud['all_rows']
    product_col = ud['product_col']
    price_col   = ud['price_col']
    volume_col  = ud.get('volume_col', -1)
    qty_col     = ud.get('qty_col', -1)
    rate        = ud.get('rate', DEFAULT_RATE)

    msg = await update.message.reply_text('⚙️ Обрабатываю…')

    try:
        vol_header = str(headers[volume_col]) if volume_col != -1 else ''
        qty_header = str(headers[qty_col])    if qty_col   != -1 else ''

        def _parse_price(val):
            if val is None:
                return 0
            # If already a number
            if isinstance(val, (int, float)):
                return int(val) if val > 0 else 0
            s = str(val).strip()
            # Strip all non-digit characters except period and comma
            s = re.sub(r'[^\d.,]', '', s)
            if not s:
                return 0
            # Handle European format: 12.000,50 → remove dots, replace comma with dot
            if ',' in s and '.' in s:
                if s.index('.') < s.index(','):
                    s = s.replace('.', '').replace(',', '.')
                else:
                    s = s.replace(',', '')
            elif ',' in s:
                # Could be thousands separator (12,000) or decimal (12,50)
                parts = s.split(',')
                if len(parts) == 2 and len(parts[1]) == 3:
                    s = s.replace(',', '')   # thousands separator
                else:
                    s = s.replace(',', '.')  # decimal comma
            try:
                return int(float(s))
            except (ValueError, TypeError):
                return 0

        data_rows = []
        skipped_no_product = 0
        skipped_no_price = 0
        for row in all_rows:
            if len(row) <= max(product_col, price_col): continue
            product = str(row[product_col] or '').strip()
            price_krw = _parse_price(row[price_col])

            if not product:
                skipped_no_product += 1
                continue
            if price_krw <= 0:
                skipped_no_price += 1
                continue

            specs = []
            if volume_col != -1 and volume_col < len(row):
                v = _format_spec(row[volume_col] or '', vol_header)
                if v: specs.append(v)
            if qty_col != -1 and qty_col < len(row):
                q = _format_spec(row[qty_col] or '', qty_header)
                if q: specs.append(f'{q}개/박스')
            if specs:
                product = f'{product} ({", ".join(specs)})'
            data_rows.append((product, price_krw))

        if not data_rows:
            # Sample a few rows to help diagnose
            sample = []
            for row in all_rows[:5]:
                if len(row) > max(product_col, price_col):
                    sample.append(f'  товар={repr(str(row[product_col] or "")[:30])}  цена={repr(str(row[price_col] or "")[:20])}')
            sample_text = '\n'.join(sample) if sample else '  (нет строк)'
            await msg.edit_text(
                f'⚠️ Не смог разобрать данные с выбранными столбцами.\n\n'
                f'Примеры значений из первых строк:\n{sample_text}\n\n'
                f'Выбери столбцы вручную — нажми кнопку ниже:'
            )
            # Reset columns so _advance_flow asks again
            ud['price_col'] = None
            ud['product_col'] = None
            ud['rate'] = None
            return await _advance_flow(update, context)

        date_str = update.message.date.strftime('%Y-%m-%d')

        # 1. Save to Google Sheets
        await msg.edit_text('💾 Сохраняю в базу…')
        gc = get_gspread_client()
        spreadsheet = gc.open_by_key(SHEET_ID)
        added = save_to_sheet(data_rows, supplier, brand, date_str, rate, spreadsheet)

        # 2. Upload to Google Drive
        await msg.edit_text('☁️ Загружаю на Drive…')
        drive_url = upload_to_drive(file_bytes, file_name, supplier, date_str)

        # 3. Forward to archive Telegram channel
        if CHANNEL_ID:
            try:
                await context.bot.send_document(
                    chat_id=CHANNEL_ID,
                    document=file_id,
                    caption=f'📦 {supplier} | {date_str} | {added} поз. | курс {rate}₩'
                )
            except Exception as e:
                logger.warning('Channel forward failed: %s', e)

        sheets_url = f'https://docs.google.com/spreadsheets/d/{SHEET_ID}'
        drive_line = (f'\n☁️ <a href="{drive_url}">Файл в Drive (папка {_esc(supplier)})</a>'
                      if drive_url else '\n⚠️ Drive: не загружено')

        brand_line = f'\n🏷️ Бренд: <b>{_esc(brand)}</b>' if brand else ''
        await msg.edit_text(
            f'✅ <b>Готово!</b>\n\n'
            f'📦 Поставщик: <b>{_esc(supplier)}</b>'
            f'{brand_line}\n'
            f'📅 Дата: {date_str}\n'
            f'💱 Курс: {rate}₩ = $1\n'
            f'➕ Добавлено в историю: <b>{added}</b> позиций\n\n'
            f'📊 <a href="{sheets_url}">Открыть таблицу базы данных</a>\n'
            f'📊 Листы обновлены: Сравнение, Лучшие цены, Бренды, Поставщики, Дашборд'
            f'{drive_line}',
            parse_mode='HTML', disable_web_page_preview=True
        )

    except Exception as e:
        logger.error('_do_process error', exc_info=True)
        await msg.edit_text(f'❌ Ошибка: {e}')

    context.user_data.clear()
    return ConversationHandler.END


# ── Search handler ─────────────────────────────────────────────────────────────
def _split_spec(name):
    m = re.match(r'^(.+?)\s*\(([^（）()]{1,30})\)\s*$', name)
    return (m.group(1).strip(), m.group(2).strip()) if m else (name, '')


async def handle_query(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text or len(text) < 2:
        return

    try:
        gc = get_gspread_client()
        sp = gc.open_by_key(SHEET_ID)
        try:
            sheet = sp.worksheet('База')
        except Exception:
            sheet = sp.sheet1
        all_data = sheet.get_all_values()
    except Exception as e:
        await update.message.reply_text(f'❌ Ошибка чтения базы: {e}')
        return

    if len(all_data) <= 1:
        await update.message.reply_text('База пустая — загрузи файлы поставщиков.')
        return

    text_lower = text.lower()
    latest = get_latest_prices(all_data)

    products = {}
    for (product, supplier), (date, price_krw, brand) in latest.items():
        if (text_lower not in product.lower()
                and text_lower not in supplier.lower()
                and text_lower not in brand.lower()):
            continue
        base_name, spec = _split_spec(product)
        products.setdefault(base_name, []).append((supplier, price_krw, spec, date, brand))

    if not products:
        await update.message.reply_text(
            f'Ничего не найдено по «{_esc(text)}».\n'
            f'Попробуй часть названия бренда, товара или поставщика.'
        )
        return

    lines = [f'🔍 <b>{_esc(text)}</b> — {len(products)} позиций\n']
    shown = 0
    for base_name in sorted(products.keys()):
        if shown >= 20:
            lines.append(f'<i>...и ещё {len(products)-shown} товаров. Уточни запрос.</i>')
            break
        entries = sorted(products[base_name], key=lambda x: x[1])
        # brand from the cheapest entry (all should be same brand for same product)
        brand_label = entries[0][4] if entries[0][4] else ''
        brand_str = f' <i>({_esc(brand_label)})</i>' if brand_label else ''
        lines.append(f'<b>{_esc(base_name)}</b>{brand_str}')
        for i, (supplier, price_krw, spec, date, _brand) in enumerate(entries):
            price_usd = round(price_krw / DEFAULT_RATE, 2)
            mark = ' ✅' if i == 0 else ''
            spec_str = f'  <i>{_esc(spec)}</i>' if spec else ''
            lines.append(
                f'  🏭 {_esc(supplier)}: <b>{price_krw:,}₩</b> / ${price_usd}{mark}{spec_str}'
            )
        lines.append('')
        shown += 1

    try:
        await update.message.reply_text('\n'.join(lines), parse_mode='HTML')
    except Exception:
        plain = re.sub(r'<[^>]+>', '', '\n'.join(lines))
        await update.message.reply_text(plain)


# ── Entry point ────────────────────────────────────────────────────────────────
def main():
    app = Application.builder().token(BOT_TOKEN).build()

    conv = ConversationHandler(
        entry_points=[MessageHandler(filters.Document.ALL, handle_document)],
        states={
            WAITING_SUPPLIER:  [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_supplier_input)],
            WAITING_BRAND:     [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_brand_input)],
            WAITING_PRICE_COL: [CallbackQueryHandler(_col_callback, pattern=r'^col:')],
            WAITING_NAME_COL:  [CallbackQueryHandler(_col_callback, pattern=r'^col:')],
            WAITING_RATE:      [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_rate_input)],
        },
        fallbacks=[CommandHandler('cancel', cmd_cancel)],
        per_user=True, per_chat=True
    )

    app.add_handler(CommandHandler('start', cmd_start))
    app.add_handler(CommandHandler('cancel', cmd_cancel))
    app.add_handler(conv)
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_query))

    port = int(os.environ.get('PORT', 8080))
    threading.Thread(
        target=lambda: HTTPServer(('0.0.0.0', port), HealthHandler).serve_forever(),
        daemon=True
    ).start()

    logger.info('Bot started')
    app.run_polling()


if __name__ == '__main__':
    main()
